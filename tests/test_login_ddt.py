import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

wb = load_workbook("data/login_data.xlsx")
sheet = wb.active

driver = webdriver.Chrome()
driver.maximize_window()

for row in range(2, sheet.max_row + 1):
    email = sheet.cell(row, 1).value
    password = sheet.cell(row, 2).value
    expected = sheet.cell(row, 3).value

    driver.get("https://www.phptravels.net/login")
    time.sleep(2)

    if email:
        driver.find_element(By.NAME, "email").send_keys(email)
    if password:
        driver.find_element(By.NAME, "password").send_keys(password)

    driver.find_element(By.ID, "submitBTN").click()
    time.sleep(3)

    if "/dashboard" in driver.current_url:
        actual = "SUCCESS"
    else:
        actual = "FAILURE"

    print(f"Row {row-1}: {actual} → {'PASSED' if actual == expected else 'FAILED'}")

    driver.delete_all_cookies()

driver.quit()